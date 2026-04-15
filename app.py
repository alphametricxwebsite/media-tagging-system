"""
Agentic AI Tagging System for Media Intelligence
Obsidian Ink Theme · Crisp Modern Design
Split-screen: Results + Chat Agent
Built with Streamlit + Claude API
"""
import streamlit as st
import json, time, re, zipfile, csv
import xml.etree.ElementTree as ET
from io import BytesIO, StringIO
from datetime import datetime

st.set_page_config(page_title="Media Intelligence Agent", page_icon="🏷️", layout="wide", initial_sidebar_state="expanded")

# ============================================================
# CLEAN LIGHT DASHBOARD — SOFT CARDS, PASTEL ACCENTS
# ============================================================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap');

html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

.stApp { background: #f2f3f7; }
header[data-testid="stHeader"] { background: #f2f3f7 !important; }
.block-container { padding-top: 1rem !important; max-width: 100% !important; }

/* Sidebar */
section[data-testid="stSidebar"] {
    background: #ffffff !important;
    border-right: 1px solid #e8e9ee !important;
    box-shadow: 2px 0 12px rgba(0,0,0,0.03) !important;
}
section[data-testid="stSidebar"] * { color: #4a5568 !important; }
section[data-testid="stSidebar"] h1, section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3, section[data-testid="stSidebar"] strong { color: #1a202c !important; }

#MainMenu, footer { visibility: hidden; }

/* === HEADER === */
.main-header {
    background: #ffffff;
    border: 1px solid #e8e9ee;
    border-radius: 20px;
    padding: 1.1rem 1.8rem;
    margin-bottom: 1rem;
    display: flex; align-items: center; gap: 1rem;
    box-shadow: 0 2px 12px rgba(0,0,0,0.04);
}
.main-header .icon { font-size: 1.6rem; background: linear-gradient(135deg, #667eea, #764ba2); -webkit-background-clip: text; -webkit-text-fill-color: transparent; }
.main-header h1 { color: #1a202c; font-size: 1.3rem; font-weight: 700; margin: 0; }
.main-header p { color: #8892a4; font-size: 0.8rem; margin: 0.1rem 0 0 0; }

/* === STATUS BAR === */
.status-bar {
    background: #ffffff;
    border: 1px solid #e8e9ee;
    border-radius: 14px;
    padding: 0.55rem 1.1rem;
    margin-bottom: 0.8rem;
    display: flex; align-items: center; gap: 0.7rem;
    box-shadow: 0 1px 6px rgba(0,0,0,0.03);
}
.status-dot { width: 8px; height: 8px; border-radius: 50%; }
.status-dot-done { background: #48bb78; box-shadow: 0 0 6px rgba(72,187,120,0.4); }
.status-dot-processing { background: #ecc94b; box-shadow: 0 0 6px rgba(236,201,75,0.4); animation: pulse 1.5s ease-in-out infinite; }
.status-dot-error { background: #fc8181; box-shadow: 0 0 6px rgba(252,129,129,0.4); }
@keyframes pulse { 0%,100%{opacity:1;} 50%{opacity:0.4;} }
.status-text { font-size: 0.8rem; font-weight: 600; color: #2d3748; }
.status-detail { font-size: 0.72rem; color: #a0aec0; margin-left: auto; font-family: 'JetBrains Mono', monospace; }

/* === STAT CHIPS === */
.stat-chip {
    display: inline-flex; align-items: center; gap: 0.35rem;
    background: linear-gradient(135deg, rgba(255,235,245,0.6), rgba(235,245,255,0.6));
    border: 1px solid rgba(255,255,255,0.8);
    border-radius: 14px; padding: 0.5rem 0.9rem;
    box-shadow: 0 1px 6px rgba(0,0,0,0.03);
    backdrop-filter: blur(8px);
}
.stat-chip .stat-num { font-size: 1.15rem; font-weight: 800; color: #1a202c; }
.stat-chip .stat-label { font-size: 0.62rem; color: #a0aec0; font-weight: 500; text-transform: uppercase; letter-spacing: 0.05em; }

/* === ARTICLE CARDS — soft pastel gradient glass === */
.article-card {
    background: linear-gradient(135deg, rgba(255,230,240,0.5) 0%, rgba(230,240,255,0.5) 40%, rgba(220,245,255,0.5) 100%);
    backdrop-filter: blur(12px);
    -webkit-backdrop-filter: blur(12px);
    border: 1px solid rgba(255,255,255,0.8);
    border-radius: 18px;
    padding: 0.9rem 1.1rem;
    margin-bottom: 0.55rem;
    transition: all 0.25s ease;
    position: relative;
    box-shadow: 0 2px 12px rgba(0,0,0,0.04);
}
.article-card:hover {
    box-shadow: 0 8px 28px rgba(0,0,0,0.08);
    transform: translateY(-2px);
    border-color: rgba(200,180,220,0.5);
}
.article-card .card-num {
    position: absolute; top: 0.5rem; right: 0.7rem;
    font-size: 0.6rem; font-weight: 700; color: #d0d5dd;
    font-family: 'JetBrains Mono', monospace;
}
.article-card .pub-name {
    font-size: 0.68rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.05em;
    color: #667eea; margin-bottom: 0.25rem;
}
.article-card .headline { font-size: 0.88rem; font-weight: 600; color: #1a202c; margin-bottom: 0.4rem; line-height: 1.35; }
.article-card .headline a { color: #1a202c; text-decoration: none; }
.article-card .headline a:hover { color: #667eea; text-decoration: underline; }

/* === PAYWALL CARD — BOLD AMBER HIGHLIGHT === */
.article-card-paywall {
    background: #fffbf0 !important;
    border: 2px solid #f6ad55 !important;
    border-left: 5px solid #ed8936 !important;
}
.article-card-paywall::after {
    content: '🔒 PAYWALL';
    position: absolute; top: 0; right: 0;
    background: linear-gradient(135deg, #ed8936, #f6ad55);
    color: #fff; font-size: 0.58rem; font-weight: 800;
    padding: 0.2rem 0.7rem 0.2rem 0.5rem;
    border-radius: 0 16px 0 10px;
    letter-spacing: 0.06em;
    box-shadow: 0 2px 8px rgba(237,137,54,0.3);
}
.paywall-banner {
    background: #fef3cd;
    border: 1px solid #fbd38d;
    border-radius: 8px;
    padding: 0.2rem 0.6rem;
    font-size: 0.65rem; font-weight: 600; color: #c05621;
    margin-bottom: 0.3rem;
    display: inline-block;
}

/* === TAG CHIPS — soft pastel on white === */
.tag-chip {
    display: inline-block; padding: 0.18rem 0.55rem; border-radius: 8px;
    font-size: 0.63rem; font-weight: 600; margin: 0.08rem 0.08rem;
}
.tag-positive { background: #f0fff4; color: #276749; border: 1px solid #c6f6d5; }
.tag-negative { background: #fff5f5; color: #9b2c2c; border: 1px solid #fed7d7; }
.tag-neutral { background: #f7fafc; color: #718096; border: 1px solid #e2e8f0; }
.tag-topic { background: #ebf4ff; color: #3182ce; border: 1px solid #bee3f8; }
.tag-paywall { background: #fffaf0; color: #c05621; border: 1px solid #feebc8; }
.tag-translated { background: #faf5ff; color: #6b46c1; border: 1px solid #e9d8fd; }
.tag-expansion { background: #e6fffa; color: #234e52; border: 1px solid #b2f5ea; }
.tag-ceo { background: #fff5f7; color: #b83280; border: 1px solid #fed7e2; }
.tag-high { background: #f0fff4; color: #276749; border: 1px solid #c6f6d5; }
.tag-medium { background: #fffaf0; color: #c05621; border: 1px solid #feebc8; }
.tag-low { background: #fff5f5; color: #9b2c2c; border: 1px solid #fed7d7; }

/* === CHAT === */
.chat-header {
    background: #ffffff;
    border: 1px solid #e8e9ee;
    border-radius: 16px;
    padding: 0.7rem 1rem; margin-bottom: 0.5rem;
    display: flex; align-items: center; gap: 0.5rem;
    box-shadow: 0 1px 6px rgba(0,0,0,0.03);
}
.chat-header .dot { width: 8px; height: 8px; border-radius: 50%; background: #48bb78; box-shadow: 0 0 6px rgba(72,187,120,0.4); }
.chat-header h3 { margin: 0; font-size: 0.82rem; font-weight: 700; color: #1a202c; }
.chat-header span { font-size: 0.68rem; color: #a0aec0; }

[data-testid="stChatMessage"] {
    background: #ffffff !important;
    border: 1px solid #e8e9ee !important;
    border-radius: 14px !important;
    box-shadow: 0 1px 4px rgba(0,0,0,0.02) !important;
}
[data-testid="stChatMessage"] p, [data-testid="stChatMessage"] span,
[data-testid="stChatMessage"] li, [data-testid="stChatMessage"] div,
[data-testid="stChatMessage"] strong, [data-testid="stChatMessage"] em { color: #2d3748 !important; }

/* === BUTTONS === */
.stButton > button {
    background: linear-gradient(135deg, #667eea, #764ba2) !important;
    color: white !important; border: none !important;
    border-radius: 12px !important; font-weight: 600 !important;
    font-family: 'Inter', sans-serif !important;
    box-shadow: 0 3px 12px rgba(102,126,234,0.3) !important;
    transition: all 0.2s ease !important;
}
.stButton > button:hover { box-shadow: 0 5px 20px rgba(102,126,234,0.45) !important; transform: translateY(-1px) !important; }
.stDownloadButton > button {
    background: #ffffff !important;
    color: #4a5568 !important;
    border: 1px solid #e2e8f0 !important;
    border-radius: 12px !important; font-weight: 600 !important;
    font-family: 'Inter', sans-serif !important;
    box-shadow: 0 1px 4px rgba(0,0,0,0.04) !important;
}
.stDownloadButton > button:hover {
    border-color: #667eea !important; color: #667eea !important;
    box-shadow: 0 3px 12px rgba(102,126,234,0.12) !important;
}

/* === TABS === */
.stTabs [data-baseweb="tab-list"] {
    background: #ffffff;
    border-radius: 14px; padding: 0.25rem; gap: 0.15rem;
    border: 1px solid #e8e9ee;
    box-shadow: 0 1px 4px rgba(0,0,0,0.02);
}
.stTabs [data-baseweb="tab"] {
    border-radius: 10px !important; font-family: 'Inter', sans-serif !important;
    font-weight: 500 !important; font-size: 0.76rem !important; color: #8892a4 !important;
    padding: 0.4rem 0.8rem !important;
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, #667eea, #764ba2) !important;
    color: #fff !important; font-weight: 600 !important;
    box-shadow: 0 2px 8px rgba(102,126,234,0.25) !important;
}

.stRadio label {
    background: #ffffff !important;
    border: 1px solid #e2e8f0 !important;
    border-radius: 10px !important; color: #4a5568 !important;
}

[data-testid="stFileUploader"] {
    background: #ffffff;
    border: 2px dashed #d0d5dd;
    border-radius: 16px; padding: 0.5rem;
}

.stProgress > div > div > div { background: linear-gradient(90deg, #667eea, #764ba2) !important; border-radius: 6px !important; }

[data-testid="stDataFrame"] { border-radius: 14px; overflow: hidden; border: 1px solid #e8e9ee; box-shadow: 0 1px 6px rgba(0,0,0,0.03); }

/* All text */
.stApp p, .stApp span, .stApp li, .stApp div, .stApp label { color: #4a5568; }
.stMarkdown h1, .stMarkdown h2, .stMarkdown h3, .stMarkdown strong { color: #1a202c !important; }
.stCode, code, pre { color: #2d3748 !important; background: #f7fafc !important; border: 1px solid #e2e8f0 !important; }
[data-testid="stAlert"] p, [data-testid="stAlert"] div { color: #2d3748 !important; }
.stCaption, [data-testid="stCaptionContainer"] p { color: #a0aec0 !important; }
.stSpinner > div, .stSpinner > div > div { color: #667eea !important; }

::-webkit-scrollbar { width: 5px; }
::-webkit-scrollbar-track { background: transparent; }
::-webkit-scrollbar-thumb { background: #d0d5dd; border-radius: 3px; }

.upload-panel {
    background: linear-gradient(135deg, rgba(255,230,240,0.4) 0%, rgba(230,240,255,0.4) 50%, rgba(220,250,250,0.4) 100%);
    backdrop-filter: blur(12px);
    border: 1px solid rgba(255,255,255,0.8);
    border-radius: 18px; padding: 2.5rem; text-align: center;
    box-shadow: 0 2px 12px rgba(0,0,0,0.04);
}
.upload-panel .icon { font-size: 2.5rem; margin-bottom: 0.5rem; }
.upload-panel .title { font-size: 1rem; font-weight: 600; color: #1a202c; margin-bottom: 0.2rem; }
.upload-panel .sub { font-size: 0.78rem; color: #a0aec0; }
</style>
""", unsafe_allow_html=True)


# ============================================================
# PAYWALL BLOCKLIST
# ============================================================
PAYWALL_DOMAINS = {
    'bloomberg.com','wsj.com','ft.com','economist.com','barrons.com','reuters.com',
    'nytimes.com','washingtonpost.com','businessinsider.com','fortune.com','forbes.com',
    'hbr.org','cnbc.com','seekingalpha.com','tikr.com','morningstar.com','zacks.com',
    'investopedia.com','thestreet.com','marketwatch.com','ainvest.com','stockstory.org',
    'simplywall.st','bitget.com','achrnews.com','contractingbusiness.com','hpac.com',
    'apnews.com','afp.com','ad-hoc-news.de','handelsblatt.com',
}

def is_paywall_url(url):
    if not url: return False
    url_lower = url.lower()
    for domain in PAYWALL_DOMAINS:
        if domain in url_lower: return True
    return False


# ============================================================
# PUBLICATION NAME NORMALIZATION (Fix #6)
# ============================================================
KNOWN_PUBLICATIONS = {
    'bloomberg': 'Bloomberg', 'reuters': 'Reuters', 'cnbc': 'CNBC',
    'the wall street journal': 'The Wall Street Journal', 'wsj': 'The Wall Street Journal',
    'the new york times': 'The New York Times', 'nyt': 'The New York Times',
    'the washington post': 'The Washington Post', 'forbes': 'Forbes',
    'fortune': 'Fortune', 'business insider': 'Business Insider',
    'the economist': 'The Economist', 'financial times': 'Financial Times',
    'associated press': 'Associated Press', 'ap news': 'AP News',
    'yahoo finance': 'Yahoo Finance', 'yahoo! finance': 'Yahoo Finance',
    'seeking alpha': 'Seeking Alpha', 'marketwatch': 'MarketWatch',
    'barron\'s': "Barron's", 'barrons': "Barron's",
    'morningstar': 'Morningstar', 'investopedia': 'Investopedia',
    'the motley fool': 'The Motley Fool', 'motley fool': 'The Motley Fool',
    'pr newswire': 'PR Newswire', 'business wire': 'Business Wire',
    'globe newswire': 'GlobeNewsWire', 'globenewswire': 'GlobeNewsWire',
    'achr news': 'ACHR News', 'achrnews': 'ACHR News',
    'contracting business': 'Contracting Business',
    'hpac engineering': 'HPAC Engineering', 'hpac': 'HPAC Engineering',
    'the air conditioning, heating & refrigeration news': 'ACHR News',
    'hvac insider': 'HVAC Insider', 'supply house times': 'Supply House Times',
    'cooling post': 'Cooling Post', 'the cooling post': 'Cooling Post',
    'zacks investment research': 'Zacks Investment Research', 'zacks': 'Zacks Investment Research',
    'stock story': 'StockStory', 'stockstory': 'StockStory',
    'ainvest': 'AInvest', 'simply wall st': 'Simply Wall St',
    'handelsblatt': 'Handelsblatt', 'afp': 'AFP',
}

def normalize_publication_name(pub):
    """Convert ALL CAPS pub names to proper case, using known mappings where possible."""
    if not pub: return pub
    pub_stripped = pub.strip()
    # Check known mappings (case-insensitive)
    lookup = pub_stripped.lower()
    if lookup in KNOWN_PUBLICATIONS:
        return KNOWN_PUBLICATIONS[lookup]
    # If it's ALL CAPS, convert to title case
    if pub_stripped == pub_stripped.upper() and len(pub_stripped) > 3:
        return pub_stripped.title()
    return pub_stripped


# ============================================================
# DOCX PARSER (unchanged logic)
# ============================================================
def parse_docx(file_bytes, filename=""):
    articles = []
    monitor_date = extract_monitor_date(filename)
    with zipfile.ZipFile(BytesIO(file_bytes)) as z:
        rels_xml = z.read("word/_rels/document.xml.rels")
        rels_root = ET.fromstring(rels_xml)
        hyperlinks = {}
        for rel in rels_root:
            if 'hyperlink' in rel.attrib.get('Type', '').lower():
                hyperlinks[rel.attrib.get('Id', '')] = rel.attrib.get('Target', '')
        doc_xml = z.read("word/document.xml")
        root = ET.fromstring(doc_xml)
        ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main',
              'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
        tables = root.findall('.//w:tbl', ns)
        if tables and len(tables[0].findall('.//w:tr', ns)) > 3:
            # TABLE-BASED FORMAT (e.g., Trane daily monitor)
            articles = _parse_table_format(tables[0], hyperlinks, ns, monitor_date)
        else:
            # PARAGRAPH-BASED FORMAT (e.g., pharma daily news summary)
            body = root.find('.//w:body', ns)
            if body is not None:
                articles = _parse_paragraph_format(body, hyperlinks, ns, monitor_date)
    return articles, monitor_date


def _parse_table_format(table, hyperlinks, ns, monitor_date):
    """Parse table-based docx format (e.g., Trane daily monitor)."""
    articles = []
    rows = table.findall('.//w:tr', ns)
    current_section = "General"
    for ri, row in enumerate(rows):
        cells = row.findall('.//w:tc', ns)
        if not cells: continue
        paragraphs = cells[0].findall('.//w:p', ns)
        para_data = []
        # Track field code state ACROSS paragraphs (fldChar begin/end can span P1→P2)
        in_field = False
        field_url = ''
        field_text_parts = []
        field_start_para = -1
        for pi, p in enumerate(paragraphs):
            parts, hl_list = [], []
            # Method 1: Standard w:hyperlink elements
            for hl in p.findall('./w:hyperlink', ns):
                rid = hl.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id', '')
                hl_t = ''.join(t.text for r in hl.findall('.//w:r', ns) for t in r.findall('.//w:t', ns) if t.text).strip()
                url = hyperlinks.get(rid, '')
                if hl_t or url: hl_list.append({'text': hl_t, 'url': url}); parts.append(hl_t)
            # Method 2: Field code hyperlinks (fldChar + instrText)
            runs = p.findall('./w:r', ns)
            for r in runs:
                fld = r.find('.//w:fldChar', ns)
                instr = r.find('.//w:instrText', ns)
                if fld is not None:
                    fld_type = ''
                    for attr_key in fld.attrib:
                        if 'fldCharType' in attr_key:
                            fld_type = fld.attrib[attr_key]
                    if fld_type == 'begin':
                        in_field = True; field_url = ''; field_text_parts = []; field_start_para = pi
                    elif fld_type == 'separate':
                        pass
                    elif fld_type == 'end':
                        if field_url:
                            ft = ''.join(field_text_parts).strip()
                            if ft or field_url:
                                if field_start_para == pi:
                                    hl_list.append({'text': ft, 'url': field_url})
                                    parts.append(ft)
                                elif field_start_para >= 0 and field_start_para < len(para_data):
                                    para_data[field_start_para]['hyperlinks'].append({'text': ft, 'url': field_url})
                                    para_data[field_start_para]['text'] = (para_data[field_start_para]['text'] + ' ' + ft).strip()
                        in_field = False; field_url = ''; field_text_parts = []; field_start_para = -1
                elif instr is not None and instr.text:
                    m = re.search(r'HYPERLINK\s+"([^"]+)"', instr.text)
                    if m: field_url = m.group(1)
                elif in_field and field_url:
                    for t in r.findall('.//w:t', ns):
                        if t.text: field_text_parts.append(t.text)
                else:
                    if not in_field:
                        for t in r.findall('.//w:t', ns):
                            if t.text: parts.append(t.text)
            para_data.append({'text': ''.join(parts).strip(), 'hyperlinks': hl_list})
        full = ' '.join(p['text'] for p in para_data).strip()
        fl = full.lower()
        if ri < 3: continue
        # Detect section headers (generic — works for any document)
        if 'benchmark companies' in fl or 'benchmark' == fl: current_section = "Benchmark Companies"; continue
        if 'trends' in fl and ('issues' in fl or 'government' in fl): current_section = "Trends / Issues / Government Relations"; continue
        # Generic section header: short text, all caps or title-like, no hyperlinks
        is_section_header = (len(full) < 80 and not any(pd.get('hyperlinks') for pd in para_data)
                            and (full.upper() == full or full.istitle()) and 'summary' not in fl)
        if is_section_header and len(para_data) <= 2 and fl not in ('', 'summary:', 'translated from'):
            current_section = full.strip()
            continue
        if not full or len(full) < 10: continue
        article = parse_article_cell(para_data, current_section, monitor_date)
        if article:
            articles.append(article)
            for exp in parse_expansion_pubs(para_data):
                exp.update({'section': current_section, 'monitor_date': monitor_date,
                           'parent_headline': article['headline'], 'parent_summary': article.get('summary','')})
                # Use parent headline for expansion articles that have no headline of their own
                if not exp.get('headline'):
                    exp['headline'] = article['headline']
                articles.append(exp)
    return articles

def extract_monitor_date(fn):
    """Extract monitor date from filename. Tries MM.DD.YY, MM.DD.YYYY, MM-DD-YY, etc.
    Also tries 'April 14, 2026' style from filename."""
    names_num = {'01':'Jan.','02':'Feb.','03':'Mar.','04':'Apr.','05':'May','06':'Jun.',
             '07':'Jul.','08':'Aug.','09':'Sep.','10':'Oct.','11':'Nov.','12':'Dec.'}
    names_word = {'january':'Jan.','february':'Feb.','march':'Mar.','april':'Apr.','may':'May',
                  'june':'Jun.','july':'Jul.','august':'Aug.','september':'Sep.','october':'Oct.',
                  'november':'Nov.','december':'Dec.'}
    # Try numeric: MM.DD.YY or MM.DD.YYYY
    for pat in [r'(\d{1,2})[._-](\d{1,2})[._-](\d{4})', r'(\d{1,2})[._-](\d{1,2})[._-](\d{2})']:
        m = re.search(pat, fn)
        if m:
            g1, g2 = m.group(1).zfill(2), m.group(2).zfill(2)
            month_str = names_num.get(g1, g1)
            day = int(g2)
            return f"{month_str} {day}"
    # Try word-based: "April 14, 2026" or "April_14__2026"
    m = re.search(r'(January|February|March|April|May|June|July|August|September|October|November|December)[_\s,]+(\d{1,2})', fn, re.IGNORECASE)
    if m:
        month_str = names_word.get(m.group(1).lower(), m.group(1)[:3] + '.')
        day = int(m.group(2))
        return f"{month_str} {day}"
    return ""

def parse_article_cell(para_data, section, monitor_date):
    if len(para_data) < 2: return None
    pub, hl, hl_url, summary, is_trans = "", "", "", "", False
    state = "pub"
    for p in para_data:
        text, hls = p['text'].strip(), p['hyperlinks']
        if not text: continue
        if text.lower().startswith('translated from'): is_trans = True; continue
        if 'also covered by' in text.lower(): continue
        if state == "pub":
            if not hls and text.upper() == text and len(text) < 100: pub = text; state = "headline"
            elif hls: pub = ""; hl = hls[0]['text']; hl_url = hls[0]['url']; state = "summary"
            else: pub = text; state = "headline"
        elif state == "headline":
            if hls: hl = hls[0]['text']; hl_url = hls[0]['url']
            else: hl = text
            state = "summary"
        elif state == "summary":
            if text.lower().startswith('summary:'): summary = text[8:].strip()
            elif text.lower() == 'summary:': continue
            elif not summary and not text.lower().startswith('this story'): summary += (" "+text) if summary else text
    if not pub and not hl: return None
    # Fix #2: For paywall articles (no URL), still keep the headline text from the docx
    # Fix #3: is_paywall is determined by blocklist later, not by absence of URL
    # Some articles legitimately have no hyperlink in the docx but have a headline
    # Fix #6: Normalize publication name from ALL CAPS
    pub = normalize_publication_name(pub)
    return {'publication': pub, 'headline': hl, 'url': hl_url, 'summary': summary[:1500],
            'section': section, 'monitor_date': monitor_date, 'is_translated': is_trans,
            'is_paywall': False, 'is_expansion': False}

def parse_expansion_pubs(para_data):
    exps = []
    for p in para_data:
        if 'also covered by' not in p['text'].lower(): continue
        for hl in p['hyperlinks']:
            if hl['text'] and hl['url']:
                exps.append({'publication': hl['text'], 'headline': '', 'url': hl['url'],
                            'summary': '', 'is_translated': False, 'is_paywall': False, 'is_expansion': True})
    return exps


def _extract_para_text_and_links(p, hyperlinks, ns):
    """Extract text and hyperlinks from a single paragraph element."""
    parts, hl_list = [], []
    # Standard w:hyperlink
    for hl in p.findall('./w:hyperlink', ns):
        rid = hl.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id', '')
        hl_t = ''.join(t.text for r in hl.findall('.//w:r', ns) for t in r.findall('.//w:t', ns) if t.text).strip()
        url = hyperlinks.get(rid, '')
        if hl_t or url: hl_list.append({'text': hl_t, 'url': url}); parts.append(hl_t)
    # Field code hyperlinks
    runs = p.findall('./w:r', ns)
    in_field = False; field_url = ''; field_text_parts = []
    for r in runs:
        fld = r.find('.//w:fldChar', ns)
        instr = r.find('.//w:instrText', ns)
        if fld is not None:
            fld_type = ''
            for ak in fld.attrib:
                if 'fldCharType' in ak: fld_type = fld.attrib[ak]
            if fld_type == 'begin': in_field = True; field_url = ''; field_text_parts = []
            elif fld_type == 'end':
                if field_url:
                    ft = ''.join(field_text_parts).strip()
                    if ft or field_url: hl_list.append({'text': ft, 'url': field_url}); parts.append(ft)
                in_field = False; field_url = ''; field_text_parts = []
        elif instr is not None and instr.text:
            m = re.search(r'HYPERLINK\s+"([^"]+)"', instr.text)
            if m: field_url = m.group(1)
        elif in_field and field_url:
            for t in r.findall('.//w:t', ns):
                if t.text: field_text_parts.append(t.text)
        else:
            if not in_field:
                for t in r.findall('.//w:t', ns):
                    if t.text: parts.append(t.text)
    return ''.join(parts).strip(), hl_list


def _parse_paragraph_format(body, hyperlinks, ns, monitor_date):
    """Parse paragraph-based docx format (e.g., pharma daily news, generic reports).
    Detects articles by finding hyperlinked headlines followed by summary text."""
    articles = []
    paragraphs = body.findall('./w:p', ns)
    current_section = "General"
    # Known section patterns
    section_keywords = ['back to top', 'tweets of interest']

    i = 0
    while i < len(paragraphs):
        text, hls = _extract_para_text_and_links(paragraphs[i], hyperlinks, ns)
        tl = text.lower().strip()

        # Skip empty paragraphs
        if not text.strip():
            i += 1; continue

        # Skip "Back to Top" links and tweet sections
        if any(kw in tl for kw in section_keywords):
            i += 1; continue

        # Detect section headers (short, no hyperlinks, looks like a heading)
        if (len(text) < 100 and not hls and not tl.startswith('(subscription')
            and not tl.startswith('202') and not any(c.isdigit() for c in text[:4])
            and text.strip() and len(text.split()) <= 12):
            # Check if it looks like a section header (title case, or all words capitalized)
            words = text.strip().split()
            cap_words = sum(1 for w in words if w[0].isupper() or w in ('&', 'of', 'and', 'the', 'for', 'in', 'to', 'a'))
            if cap_words >= len(words) * 0.7 and len(words) >= 2:
                current_section = text.strip()
                i += 1; continue

        # Detect articles: paragraph with a hyperlink (headline link)
        if hls and hls[0].get('url') and not hls[0]['url'].startswith('#'):
            headline = hls[0]['text'] or text
            url = hls[0]['url']
            publication = ''
            date_published = ''

            # Try to extract date and publication from the NEXT paragraph
            # Pattern: "2026/04/13, Publication Name by Author..."
            # or: "2026/04/13, Publication Name"
            summary_parts = []
            also_covered = []
            j = i + 1
            while j < len(paragraphs):
                ntxt, nhls = _extract_para_text_and_links(paragraphs[j], hyperlinks, ns)
                ntl = ntxt.lower().strip()
                if not ntxt.strip():
                    j += 1; continue
                # Check for date/publication line: "2026/04/13, Source Name..."
                date_match = re.match(r'(\d{4}/\d{2}/\d{2})\s*,\s*(.+)', ntxt)
                if date_match and not publication:
                    raw_date = date_match.group(1)
                    rest = date_match.group(2).strip()
                    # Parse date
                    try:
                        dp = raw_date.split('/')
                        mn = {'01':'Jan.','02':'Feb.','03':'Mar.','04':'Apr.','05':'May','06':'Jun.',
                              '07':'Jul.','08':'Aug.','09':'Sep.','10':'Oct.','11':'Nov.','12':'Dec.'}
                        date_published = f"{mn.get(dp[1], dp[1])} {int(dp[2])}"
                    except: pass
                    # Extract publication: "Source Name by Author" or "Source Name Press Release"
                    pub_match = re.match(r'(.+?)(?:\s+by\s+|Press Release|$)', rest)
                    if pub_match:
                        publication = pub_match.group(1).strip()
                        # Clean trailing "Press Release" etc
                        publication = re.sub(r'\s*Press\s*Release\s*$', '', publication).strip()
                    j += 1; continue
                # Check for "also covered by" links
                if nhls and not ntl.startswith('202'):
                    # Could be "also covered by" links OR next article headline
                    # If it looks like a new article (has a long headline-like hyperlink), break
                    if len(nhls[0].get('text', '')) > 30 and nhls[0].get('url', '') and 'back to top' not in ntl:
                        break
                    for nh in nhls:
                        if nh.get('url') and nh.get('text') and 'back to top' not in nh['text'].lower():
                            also_covered.append(nh)
                    j += 1; continue
                # Summary text (no links, not a date line, not a section header)
                if not nhls and not ntl.startswith('202') and len(ntxt) > 30:
                    summary_parts.append(ntxt)
                    j += 1; continue
                # Unknown — probably next article or section, break
                break

            summary = ' '.join(summary_parts)[:1500]
            pub = normalize_publication_name(publication) if publication else ''

            # Check subscription required
            is_paywall = '(subscription required)' in text.lower() or '(subscription required)' in summary.lower()

            article = {
                'publication': pub, 'headline': headline, 'url': url,
                'summary': summary, 'section': current_section,
                'monitor_date': monitor_date, 'date_published': date_published,
                'is_translated': False, 'is_paywall': is_paywall, 'is_expansion': False
            }
            articles.append(article)

            # Add "also covered by" as expansion articles
            for ac in also_covered:
                if ac.get('text') and ac.get('url') and 'back to top' not in ac['text'].lower():
                    exp = {
                        'publication': normalize_publication_name(ac['text']),
                        'headline': headline, 'url': ac['url'],
                        'summary': summary[:500], 'section': current_section,
                        'monitor_date': monitor_date, 'date_published': date_published,
                        'parent_headline': headline, 'parent_summary': summary[:500],
                        'is_translated': False, 'is_paywall': False, 'is_expansion': True
                    }
                    articles.append(exp)

            i = j; continue

        i += 1

    return articles


# ============================================================
# TAGGER (parallel batching, confidence, Haiku fallback)
# ============================================================
def safe_extract_text(response):
    text = ""
    if not hasattr(response, 'content'): return text
    for block in response.content:
        bt = getattr(block, 'text', None)
        if bt is not None and isinstance(bt, str): text += bt
    return text

def _process_single_batch(client, batch, batch_num, total_batches, analyst_examples):
    sp = build_system_prompt(analyst_examples)
    # Enable web_search for: articles with URLs to read, AND articles without URLs that need searching
    has_url_articles = any(a.get('url') and not is_paywall_url(a.get('url','')) for a in batch)
    has_no_url_articles = any(not a.get('url') and not a.get('is_paywall') for a in batch)
    needs_web = has_url_articles or has_no_url_articles
    result = {'batch_num': batch_num, 'tagged': [], 'ok': False}
    try:
        atxt = build_articles_text(batch, summary_only=False)
        up = build_user_prompt(len(batch), atxt)
        kwargs = dict(model="claude-sonnet-4-20250514", max_tokens=4096, system=sp, messages=[{"role":"user","content":up}])
        if needs_web: kwargs['tools'] = [{"type":"web_search_20250305","name":"web_search"}]
        r = client.messages.create(**kwargs)
        parsed = extract_json(safe_extract_text(r))
        if parsed: merge_results(result['tagged'], batch, parsed); result['ok'] = True; return result
    except Exception: pass
    try:
        atxt2 = build_articles_text(batch, summary_only=True)
        up2 = build_user_prompt(len(batch), atxt2)
        r2 = client.messages.create(model="claude-haiku-4-5-20251001", max_tokens=4096, system=sp, messages=[{"role":"user","content":up2}])
        parsed2 = extract_json(safe_extract_text(r2))
        if parsed2: merge_results(result['tagged'], batch, parsed2); result['ok'] = True; return result
    except Exception: pass
    for a in batch: a['tags_failed'] = True; result['tagged'].append(a)
    return result

def tag_articles_batch(articles, api_key, analyst_examples="", progress_callback=None):
    import anthropic
    from concurrent.futures import ThreadPoolExecutor, as_completed
    client = anthropic.Anthropic(api_key=api_key)
    BATCH_SIZE, MAX_PARALLEL = 8, 3
    tagged, batches = [], [articles[i:i+BATCH_SIZE] for i in range(0, len(articles), BATCH_SIZE)]
    total_batches = len(batches)
    if progress_callback: progress_callback(0.05, f"Processing {len(articles)} articles in {total_batches} batches...")
    ok, fail, completed = 0, 0, 0
    with ThreadPoolExecutor(max_workers=MAX_PARALLEL) as executor:
        futures = {executor.submit(_process_single_batch, client, b, bn, total_batches, analyst_examples): bn for bn, b in enumerate(batches, 1)}
        for future in as_completed(futures):
            completed += 1; pct = 0.05 + (completed / total_batches) * 0.90
            try:
                result = future.result(); tagged.extend(result['tagged'])
                if result['ok']: ok += 1
                else: fail += 1
                if progress_callback: progress_callback(pct, f"Batch {completed}/{total_batches} done ({ok} OK)")
            except Exception:
                fail += 1
                if progress_callback: progress_callback(pct, f"Batch error — {completed}/{total_batches}")
    if progress_callback: progress_callback(1.0, f"Done! {ok}/{total_batches} batches OK.")
    # Post-processing: fix common AI output issues
    tagged = post_process_tagged(tagged)
    return tagged


def post_process_tagged(tagged):
    """Fix common issues in AI-tagged output."""
    from datetime import datetime
    today_dd_mon = datetime.now().strftime('%-d-%b')  # e.g., "15-Apr"
    today_mon_dd = datetime.now().strftime('%b. %-d')  # e.g., "Apr. 15"
    today_full = datetime.now().strftime('%B %-d')  # e.g., "April 15"

    seen_keys = set()
    deduped = []
    for art in tagged:
        # Fix None/missing sentiment — default to Neutral (per feedback: NEVER default to Positive)
        brands = art.get('brands', {})
        for bname, bdata in brands.items():
            if isinstance(bdata, dict):
                sent = bdata.get('sentiment', '')
                if not sent or sent == 'None' or sent not in ('Positive', 'Negative', 'Neutral'):
                    bdata['sentiment'] = 'Neutral'

        # Fix date_published — reject today's date (AI hallucination), leave empty if can't determine
        dp = str(art.get('date_published', '') or '').strip()
        if dp in ('None', '', today_dd_mon, today_mon_dd, today_full):
            art['date_published'] = ''  # Leave empty — better than wrong date

        # Fix CEO — ensure empty string is default, not some hallucinated value
        ceo = str(art.get('ceo_mention', '') or '').strip()
        if ceo.lower() in ('none', 'n/a', 'not mentioned', 'no', 'null'):
            art['ceo_mention'] = ''

        # Normalize publication name
        pub = art.get('publication', '')
        if pub:
            art['publication'] = normalize_publication_name(pub)
        # Reject generic publication names
        pub_lower = (art.get('publication', '') or '').strip().lower()
        if pub_lower in ('press release', 'news article', 'article', 'unknown'):
            art['publication_flagged'] = True

        # Fix article_or_pr — ensure it's never empty
        if not art.get('article_or_pr') or art['article_or_pr'] == 'None':
            art['article_or_pr'] = 'Article'

        # Deduplicate exact duplicate entries (same headline + pub + url)
        hl = (art.get('headline', '') or '').strip().lower()
        pub_key = (art.get('publication', '') or '').strip().lower()
        url_key = (art.get('url', '') or '').strip().lower()
        dedup_key = f"{pub_key}|||{hl}|||{url_key}"
        if dedup_key in seen_keys and hl:
            continue
        if hl:
            seen_keys.add(dedup_key)

        deduped.append(art)

    return deduped

def build_articles_text(batch, summary_only=False):
    text = ""
    for j, art in enumerate(batch):
        md = art.get('monitor_date', '')
        text += f"\n--- ARTICLE {j+1} ---\nPublication: {art['publication']}\nHeadline: {art['headline']}\nURL: {art.get('url','NONE')}\nSection: {art['section']}\nMonitor Date: {md}\n"
        url = art.get('url',''); blocked = is_paywall_url(url)
        if summary_only or blocked:
            # Truly paywalled — use summary only
            s = art.get('summary','') or art.get('parent_summary','')
            text += f"PAYWALLED - Tag from summary only:\n{s[:800]}\n"
            art['is_paywall'] = True
        elif not url:
            # No URL in docx but NOT paywalled — search for it by headline
            s = art.get('summary','') or art.get('parent_summary','')
            text += f"NO URL IN DOCUMENT - Search for this article by headline to find the actual URL and publication date. Use web_search with the headline.\n"
            text += f"Summary for context: {s[:600]}\n"
            text += f"IMPORTANT: Return the actual article URL in the 'url' field so we can hyperlink the headline.\n"
        else:
            # Has URL — read the full article
            text += f"HAS URL - Read full article via web search. Extract the actual article publication date.\n"
            if art.get('summary'): text += f"Backup summary: {art['summary'][:500]}\n"
    return text

def build_user_prompt(count, atxt):
    return f"""Tag these {count} articles. For each article with a URL, use web_search to read the full article. For articles without URLs, search by headline.

{atxt}

Respond ONLY with a valid JSON array. No markdown, no commentary. Each item:
{{"publication":"Actual Source Name","headline":"Exact Original Headline","url":"...","date_published":"DD-Mon","article_or_pr":"Article",
"brands":{{"Trane Technologies":{{"sentiment":"Positive/Negative/Neutral","mentioned":true/false}},"Carrier":{{"sentiment":"...","mentioned":true/false}},"Honeywell":{{"sentiment":"...","mentioned":true/false}},"JCI":{{"sentiment":"...","mentioned":true/false}},"Daikin":{{"sentiment":"...","mentioned":true/false}},"Lennox":{{"sentiment":"...","mentioned":true/false}}}},
"topics":{{"Sustainability":true/false,"Decarbonization":true/false,"Innovation":true/false,"Energy Efficiency":true/false,"Digitization":true/false,"Electrification":true/false,"Workforce Development":true/false,"Financial Performance":true/false,"No Category Match":true/false}},
"confidence":{{"sentiment":"high/medium/low","topics":"high/medium/low","overall":"high/medium/low"}},
"ceo_mention":"","is_paywall":true/false,"publication_flagged":true/false}}

MANDATORY VALIDATION BEFORE OUTPUT — check EVERY row:
1. publication = actual source name (NEVER "Press Release", "News Article", "Ad Hoc News" if real source exists inside)
2. headline = EXACT original headline from the article (no rewriting, no summarizing, no merging)
3. date_published = extracted from the article page (format: DD-Mon e.g. "13-Apr"). NEVER use monitor date or Word file date.
4. sentiment = brand-specific, defaults to Neutral when unclear. Stock mention alone ≠ Positive.
5. ceo_mention = ONLY for tracked brand CEOs. Empty string if no CEO present. NEVER default to CEO mention.
6. topics = tag ONLY clearly relevant categories. Do NOT overcode. If none match → No Category Match only.
7. Each article appears ONCE in the array.

If the same mistake appears more than once, stop and re-evaluate your tagging logic before continuing."""

def build_system_prompt(examples=""):
    base = """You are an Advanced Media Intelligence Coding Agent for AlphaMetricx.
You MUST behave like a trained human analyst — deterministic, precise, validated.
You MUST NOT default, assume, or guess. Every field must be evidence-based.

⚠️ CRITICAL ERRORS YOU PREVIOUSLY MADE (MUST NOT REPEAT):
- Defaulting sentiment to "Positive" without evidence
- Tagging CEO mention when no CEO was referenced
- Using "Press Release" as publication name instead of actual source
- Overcoding topics (tagging 4-5 topics when only 1-2 apply)
- Using monitor date instead of actual article publication date
- Rewriting or merging headlines instead of copying the exact original

═══════════════════════════════════════
BRANDS
═══════════════════════════════════════
Tracked brands: Trane Technologies (includes METUS, Mitsubishi Electric Trane, Thermo King), Carrier, Honeywell, JCI (Johnson Controls), Daikin, Lennox.

Set "mentioned": true ONLY for brands actually named or discussed in the article.
If an article is purely about Carrier → ONLY Carrier.mentioned=true.
If an article mentions Trane AND Carrier → both mentioned=true (article appears in BOTH tabs).

═══════════════════════════════════════
PUBLICATION NAME (STRICT)
═══════════════════════════════════════
❌ NEVER use generic names: "Press Release", "News Article", "Ad Hoc News" (if actual source exists)
✅ ALWAYS extract the actual publisher/source name from the article page
Examples: Business Wire, PR Newswire, Reuters, Economic Times, Bloomberg, ACHR News
If the Word doc says "AD HOC NEWS" but the article is sourced from a real publication, use the real name.
Use proper title case — not ALL CAPS.
Set publication_flagged=true if unfamiliar or unsure of correct format.

═══════════════════════════════════════
HEADLINE (STRICT)
═══════════════════════════════════════
Extract the EXACT original headline from the article.
❌ DO NOT rewrite, summarize, truncate, or merge headlines.
❌ DO NOT create composite headlines from multiple sources.
✅ Copy the exact headline as it appears on the article page.
If duplicate articles from different publications exist, each keeps its own unique headline.

═══════════════════════════════════════
DATE PUBLISHED (STRICT)
═══════════════════════════════════════
MUST click/read the article and extract the real publication date.
Format: DD-Mon (e.g., "13-Apr", "07-Mar", "28-Feb")
❌ NEVER use the monitor date or Word file date as date_published.
❌ NEVER use today's date.
✅ Extract from the article page only.
If truly cannot determine → leave empty (post-processing will handle it).

═══════════════════════════════════════
ARTICLE vs PRESS RELEASE
═══════════════════════════════════════
Read the actual content to determine:
- "Article" = journalist-written reporting, analysis, news coverage
- "Press Release" = official company/organization announcement via PR wire (Business Wire, PR Newswire, GlobeNewsWire, ACCESS Newswire, EIN Presswire) or company newsroom
❌ DO NOT classify journalist coverage of a press release as "Press Release" — it's an "Article"
If in doubt → "Article"

═══════════════════════════════════════
SENTIMENT (MOST CRITICAL — BRAND-SPECIFIC)
═══════════════════════════════════════
Sentiment is PER BRAND, not per article. Same article → different sentiment per brand tab.

❌ NEVER default to Positive
❌ Stock mention alone ≠ Positive
👉 If unclear → ALWAYS use "Neutral"

Rules:
- Positive: Article explicitly portrays the brand favorably — awards, explicit stock RISE mentioned, product wins, CEO praise with substance
- Negative: Article explicitly portrays the brand unfavorably — lawsuits, explicit stock FALL mentioned, recalls, downgrades, "faces pressure"
- Neutral: ALL other cases — factual reporting, passing mentions, general industry coverage, stock mentioned without explicit up/down

STOCK PERFORMANCE NORMALIZATION:
- Stock rise explicitly mentioned → Positive
- Stock fall explicitly mentioned → Negative  
- Stock mentioned without explicit movement direction → Neutral
- Growth/expansion mentioned ≠ Positive (unless stock rise is explicit)

═══════════════════════════════════════
CEO / EXECUTIVE MENTION (STRICT)
═══════════════════════════════════════
ONLY for CEOs of the 6 tracked brands:
- Trane Technologies CEO (Dave Regnery)
- Carrier CEO, Honeywell CEO, JCI CEO, Daikin CEO, Lennox CEO

❌ DO NOT tag CEO if no CEO is referenced in the article
❌ DO NOT tag other companies' CEOs (AHRI, suppliers, partners, etc.)
❌ NEVER default to "CEO mention" — empty string ("") is the default

Rules:
- "CEO quote" = direct quote or clearly paraphrased statement attributed to a tracked brand CEO
  Example: "We are accelerating innovation," said CEO Dave Regnery → "CEO quote"
- "CEO mention" = CEO name referenced WITHOUT a statement
- "CEO interview" = article is an interview format with a tracked brand CEO
- "" (empty) = no tracked brand CEO mentioned — THIS IS THE DEFAULT

═══════════════════════════════════════
TOPIC CODING (SEMANTIC — NOT KEYWORD)
═══════════════════════════════════════
Map content by MEANING, not keywords. Tag with "x" (true) only if the article SUBSTANTIVELY discusses the topic.

❌ DO NOT overcode — tagging 4-5 topics blindly is wrong
❌ DO NOT tag a topic just because the company operates in that space
✅ Tag ONLY topics that the article actually discusses with substance
✅ Evaluate ALL 9 categories for every article

1. Sustainability — Climate action, ESG goals being executed, emission commitments being met
2. Decarbonization — Active carbon/emission reduction solutions being implemented
3. Innovation — New HVAC tech, AI, automation, new products being launched/deployed
4. Energy Efficiency — Reduced energy consumption, optimization systems in operation
5. Digitization — AI, IoT, smart building systems, data centers, predictive analytics deployed
6. Electrification — Electric heating replacing fossil fuels, heat pumps being manufactured/deployed
7. Workforce Development — Training programs, hiring, skills development, apprenticeships running
8. Financial Performance — Stock, revenue, earnings, investor insights, analyst ratings, M&A
9. No Category Match — ONLY if NONE of the above 8 apply. Broader industry news.

ANTI-OVERCODING RULES:
- A pure stock article → Financial Performance ONLY (not Innovation/Digitization/etc.)
- A company announcement about future plans → likely No Category Match (not actionable yet)
- An industry overview mentioning a brand in passing → No Category Match
- Only tag multiple topics if the article genuinely covers multiple themes with substance

═══════════════════════════════════════
MULTI-COVERAGE HANDLING
═══════════════════════════════════════
When "This story was also covered by [Source1], [Source2]..." appears:
- Each linked source = SEPARATE article entry
- Open each link, extract: publication, headline, date
- Code each individually (topics/sentiment may differ per source)

═══════════════════════════════════════
SELF-CORRECTION LOOP (MANDATORY)
═══════════════════════════════════════
Before outputting JSON, re-check ALL rows:
✔ Is publication the ACTUAL source name?
✔ Is date from the article page (not monitor date)?
✔ Is sentiment evidence-based (not defaulted to Positive)?
✔ Is CEO field empty when no tracked brand CEO exists?
✔ Are topics limited to what's actually discussed?
✔ Is headline the exact original?

"If the same mistake appears more than once, stop and re-evaluate your tagging logic before continuing."

ALWAYS respond with ONLY valid JSON array."""
    if examples: base += f"\nANALYST REFERENCE EXAMPLES:\n{examples}"
    return base

def merge_results(tagged, batch, parsed):
    for j, td in enumerate(parsed):
        if j < len(batch):
            orig = batch[j]
            # Start with parser data, overlay AI tags
            m = {**orig, **td}
            # ALWAYS preserve parser's authoritative metadata fields
            m['monitor_date'] = orig.get('monitor_date', '')
            m['section'] = orig.get('section', '')
            m['is_translated'] = orig.get('is_translated', False)
            m['is_expansion'] = orig.get('is_expansion', False)
            # URL: Use parser's URL if it exists (extracted from docx hyperlink).
            # If parser had NO URL, use the AI's URL (found via web_search).
            orig_url = (orig.get('url', '') or '').strip()
            ai_url = (td.get('url', '') or '').strip()
            if orig_url:
                m['url'] = orig_url
            elif ai_url and ai_url.lower() not in ('none', 'null', ''):
                m['url'] = ai_url
            else:
                m['url'] = ''
            # Headline: Prefer parser's headline (from docx), fall back to AI's
            orig_hl = (orig.get('headline', '') or orig.get('parent_headline', '') or '').strip()
            ai_hl = (td.get('headline', '') or '').strip()
            m['headline'] = orig_hl if orig_hl else ai_hl
            # Publication: Prefer parser's publication, fall back to AI's
            orig_pub = (orig.get('publication', '') or '').strip()
            ai_pub = (td.get('publication', '') or '').strip()
            m['publication'] = orig_pub if orig_pub else ai_pub
            tagged.append(m)
    for j in range(len(parsed), len(batch)): batch[j]['tags_failed'] = True; tagged.append(batch[j])

def extract_json(text):
    if not text or not text.strip(): return None
    text = re.sub(r'```json\s*','',text); text = re.sub(r'```\s*','',text); text = text.strip()
    s, e = text.find('['), text.rfind(']')
    if s >= 0 and e > s:
        try:
            r = json.loads(text[s:e+1])
            if isinstance(r, list): return r
        except json.JSONDecodeError: pass
    try:
        r = json.loads(text)
        if isinstance(r, list): return r
    except json.JSONDecodeError: pass
    return None


# ============================================================
# CHAT AGENT
# ============================================================
def process_chat_command(user_msg, tagged_articles, api_key):
    import anthropic
    client = anthropic.Anthropic(api_key=api_key)
    articles_summary = []
    for i, art in enumerate(tagged_articles):
        td = art.get('topics', {}); conf = art.get('confidence', {})
        active = [t for t, v in td.items() if v]
        articles_summary.append(f"[{i+1}] {art.get('publication','')} — \"{art.get('headline','')[:70]}\" | Topics: {', '.join(active) or 'None'} | Conf: {conf.get('overall','?')} | PW: {art.get('is_paywall',False)} | CEO: {art.get('ceo_mention','')}")
    system = f"""You are an AI assistant helping a media analyst review tagged articles.
{len(tagged_articles)} articles tagged. Current state:
{chr(10).join(articles_summary)}

Commands: CHANGE tags, QUERY articles, EXPLAIN reasoning, BULK ops, GENERATE Excel.
For changes: {{"action":"update","changes":[{{"article_index":0,"field":"topics.Sustainability","value":true}}]}}
For export: {{"action":"generate"}}
Be concise."""
    try:
        r = client.messages.create(model="claude-sonnet-4-20250514", max_tokens=2000, system=system, messages=[{"role":"user","content":user_msg}])
        return safe_extract_text(r)
    except Exception as e: return f"Error: {str(e)[:100]}"

def apply_chat_changes(tagged_articles, response_text):
    changes_made = 0
    try:
        start = response_text.find('{"action"')
        if start < 0: return 0
        depth = 0
        for i in range(start, len(response_text)):
            if response_text[i] == '{': depth += 1
            elif response_text[i] == '}': depth -= 1
            if depth == 0:
                data = json.loads(response_text[start:i+1])
                if data.get('action') == 'update':
                    for ch in data.get('changes', []):
                        idx, field, val = ch.get('article_index', -1), ch.get('field', ''), ch.get('value')
                        if 0 <= idx < len(tagged_articles) and field:
                            parts = field.split('.'); obj = tagged_articles[idx]
                            for p in parts[:-1]:
                                if p not in obj: obj[p] = {}
                                obj = obj[p]
                            obj[parts[-1]] = val; changes_made += 1
                elif data.get('action') == 'generate': return -1
                break
    except (json.JSONDecodeError, ValueError): pass
    return changes_made


# ============================================================
# EXCEL + CSV GENERATORS
# ============================================================
def generate_excel(tagged_articles, monitor_date):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    brands = ['Trane Technologies','Carrier','Honeywell','JCI','Daikin','Lennox']
    topics = ['Sustainability','Decarbonization','Innovation','Energy Efficiency','Digitization',
              'Electrification','Workforce Development','Financial Performance','No Category Match']
    headers = ['Monitor Date','Date Published','Publication','Headline','Article OR Press Release',
               'Sentiment'] + topics + ['CODED?','CEO/Executive Mention']
    hf=Font(name='Calibri',bold=True,size=10,color='FFFFFF')
    hfl=PatternFill(start_color='1e1e2a',end_color='1e1e2a',fill_type='solid')
    df=Font(name='Calibri',size=10);lf=Font(name='Calibri',size=10,color='6366F1',underline='single')
    pf=PatternFill(start_color='FFF2CC',end_color='FFF2CC',fill_type='solid')
    pub_flag_fill=PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')  # Bright yellow for flagged publications
    bd=Border(left=Side(style='thin',color='D9D9D9'),right=Side(style='thin',color='D9D9D9'),top=Side(style='thin',color='D9D9D9'),bottom=Side(style='thin',color='D9D9D9'))
    wb.remove(wb.active)
    ws_tot=wb.create_sheet("Totals");ws_tot['B2']='Honeywell';ws_tot['C2']='Carrier';ws_tot['D2']='JCI';ws_tot['E2']='Daikin';ws_tot['F2']='Lennox';ws_tot['G2']='Trane'
    for brand in brands:
        ws=wb.create_sheet(brand)
        for ci,h in enumerate(headers,1):c=ws.cell(row=1,column=ci,value=h);c.font=hf;c.fill=hfl;c.alignment=Alignment(horizontal='center',wrap_text=True);c.border=bd
        ws.column_dimensions['A'].width=12;ws.column_dimensions['B'].width=14;ws.column_dimensions['C'].width=22;ws.column_dimensions['D'].width=55;ws.column_dimensions['E'].width=18;ws.column_dimensions['F'].width=12
        for ci in range(7,16):ws.column_dimensions[chr(64+ci)].width=14
        ws.column_dimensions['P'].width=8;ws.column_dimensions['Q'].width=20
        ba=get_brand_articles(tagged_articles,brand)
        for ri,art in enumerate(ba,2):
            brd=art.get('brands',{}).get(brand,{});td=art.get('topics',{})
            ws.cell(row=ri,column=1,value=art.get('monitor_date','')).font=df;ws.cell(row=ri,column=2,value=art.get('date_published','')).font=df
            pub_cell=ws.cell(row=ri,column=3,value=art.get('publication',''))
            pub_cell.font=df
            if art.get('publication_flagged'):
                pub_cell.fill=pub_flag_fill  # Highlight flagged publications in yellow
            hc=ws.cell(row=ri,column=4,value=art.get('headline','') or '(no headline)')
            url = art.get('url','')
            if url and art.get('headline','').strip():
                hc.hyperlink=url;hc.font=lf
            else:
                hc.font=df
            ws.cell(row=ri,column=5,value=art.get('article_or_pr','Article')).font=df;ws.cell(row=ri,column=6,value=brd.get('sentiment','Neutral')).font=df
            for ti,topic in enumerate(topics):v='x' if td.get(topic,False) else '';ws.cell(row=ri,column=7+ti,value=v).font=df;ws.cell(row=ri,column=7+ti).alignment=Alignment(horizontal='center')
            ws.cell(row=ri,column=16,value='x').font=df;ws.cell(row=ri,column=16).alignment=Alignment(horizontal='center');ws.cell(row=ri,column=17,value=art.get('ceo_mention','')).font=df
            if art.get('is_paywall'):
                for ci in range(1,18):ws.cell(row=ri,column=ci).fill=pf
            for ci in range(1,18):ws.cell(row=ri,column=ci).border=bd
        ws.freeze_panes='A2'
    cm={'Honeywell':'B','Carrier':'C','JCI':'D','Daikin':'E','Lennox':'F','Trane Technologies':'G'}
    for b in brands:ws_tot[f'{cm.get(b,"B")}3']=len(get_brand_articles(tagged_articles,b))
    ws_def=wb.create_sheet("Key Topics & Definitions");ws_def['A1']='Key Topics';ws_def['B1']='Definitions'
    for di,(t,d) in enumerate([
        ('Sustainability','Climate change action — brand actively implementing sustainability initiatives or achieving milestones'),
        ('Decarbonization','Reducing carbon emissions — brand actively reducing carbon footprint or deploying low-carbon tech'),
        ('Innovation','Advancing HVAC tech — brand launching new products, deploying new technology, R&D breakthroughs'),
        ('Energy Efficiency','Optimized energy use — brand delivering measurably more efficient products/systems'),
        ('Digitization','Digital tech, AI, IoT — brand deploying digital solutions, smart building tech, connected systems'),
        ('Electrification','Electrification of heat — brand manufacturing/deploying heat pumps, electric HVAC, gas-to-electric conversion'),
        ('Workforce Development','Training, upskilling — brand running training programs, hiring initiatives, apprenticeships'),
        ('Financial Performance','Revenue, earnings, stock — all financial results, stock performance, analyst ratings, deals'),
        ('No Category Match','Broader HVAC — none of the above 8 topics apply')
    ],2):ws_def[f'A{di}']=t;ws_def[f'B{di}']=d
    out=BytesIO();wb.save(out);out.seek(0);return out

def get_brand_articles(tagged, brand):
    """Get articles for a specific brand tab. Primary: AI's brand.mentioned flag.
    Fallback: section-based assignment for articles where AI didn't set mentioned flags."""
    result=[];bl=brand.lower()
    aliases={'trane technologies':['trane','metus','mitsubishi electric trane','thermo king'],
             'carrier':['carrier global','carrier'],'honeywell':['honeywell'],
             'jci':['johnson controls','jci'],'daikin':['daikin'],'lennox':['lennox']}
    ba=aliases.get(bl,[bl])
    for art in tagged:
        bi=art.get('brands',{}).get(brand,{})
        # Primary: AI explicitly said this brand is mentioned
        if bi.get('mentioned',False):
            result.append(art)
            continue
        # Fallback: Section-based assignment ONLY if no brand has mentioned=True
        # (means the AI didn't set any brand flags, so fall back to document section)
        any_brand_mentioned = any(
            art.get('brands',{}).get(b,{}).get('mentioned',False)
            for b in ['Trane Technologies','Carrier','Honeywell','JCI','Daikin','Lennox']
        )
        if not any_brand_mentioned:
            sec=art.get('section','').lower()
            txt=(art.get('headline','')+' '+art.get('summary','')).lower()
            if bl=='trane technologies' and 'trane' in sec:
                result.append(art); continue
            if 'benchmark' in sec or 'trends' in sec:
                for a in ba:
                    if a in txt:result.append(art);break
    return result

def generate_csv(tagged, topics):
    out=StringIO();w=csv.writer(out)
    w.writerow(['#','Monitor Date','Date Published','Publication','Headline','URL','Type','Sentiment','Section','Confidence']+topics+['CODED?','CEO','Paywall','Translated'])
    for i,art in enumerate(tagged):
        td=art.get('topics',{});conf=art.get('confidence',{})
        r=[i+1,art.get('monitor_date',''),art.get('date_published',''),art.get('publication',''),art.get('headline',''),art.get('url',''),art.get('article_or_pr','Article'),'',art.get('section',''),conf.get('overall','')]
        for t in topics:r.append('x' if td.get(t,False) else '')
        r.extend(['x',art.get('ceo_mention',''),'Yes' if art.get('is_paywall') else '','Yes' if art.get('is_translated') else ''])
        w.writerow(r)
    return out.getvalue()


# ============================================================
# TEMPLATE GENERATOR (Manual Tagging — columns A-E only)
# ============================================================
def generate_template_excel(articles, monitor_date):
    """Generate Excel template with columns A-D populated (parser data).
    Columns E onwards have headers but are blank for manual analyst tagging.
    Auto-detects format: Trane-style (brand tabs) vs generic (section tabs)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    brands = ['Trane Technologies','Carrier','Honeywell','JCI','Daikin','Lennox']
    trane_keywords = ['trane','carrier','honeywell','johnson controls','jci','daikin','lennox','metus','thermo king']

    # Detect if this is a Trane-format document by checking sections/content
    all_text = ' '.join((a.get('section','')+' '+a.get('headline','')+' '+a.get('summary','')).lower() for a in articles)
    is_trane_format = any(kw in all_text for kw in trane_keywords)

    # Common headers and styles
    headers = ['Monitor Date','Date Published','Publication','Headline','Article OR Press Release',
               'Sentiment','Notes']
    hf=Font(name='Calibri',bold=True,size=10,color='FFFFFF')
    hfl=PatternFill(start_color='1e1e2a',end_color='1e1e2a',fill_type='solid')
    df=Font(name='Calibri',size=10);lf=Font(name='Calibri',size=10,color='6366F1',underline='single')
    pf=PatternFill(start_color='FFF2CC',end_color='FFF2CC',fill_type='solid')
    bd=Border(left=Side(style='thin',color='D9D9D9'),right=Side(style='thin',color='D9D9D9'),top=Side(style='thin',color='D9D9D9'),bottom=Side(style='thin',color='D9D9D9'))
    wb.remove(wb.active)

    def _write_articles_to_sheet(ws, article_list):
        """Write articles to a worksheet (columns A-D populated, rest blank)."""
        for ci,h in enumerate(headers,1):
            c=ws.cell(row=1,column=ci,value=h);c.font=hf;c.fill=hfl;c.alignment=Alignment(horizontal='center',wrap_text=True);c.border=bd
        ws.column_dimensions['A'].width=12;ws.column_dimensions['B'].width=14;ws.column_dimensions['C'].width=22
        ws.column_dimensions['D'].width=60;ws.column_dimensions['E'].width=18;ws.column_dimensions['F'].width=12;ws.column_dimensions['G'].width=30
        for ri,art in enumerate(article_list,2):
            ws.cell(row=ri,column=1,value=art.get('monitor_date','')).font=df
            ws.cell(row=ri,column=2,value=art.get('date_published','')).font=df
            ws.cell(row=ri,column=3,value=art.get('publication','')).font=df
            hc=ws.cell(row=ri,column=4,value=art.get('headline','') or '(no headline)')
            url = art.get('url','')
            if url and art.get('headline','').strip():
                hc.hyperlink=url;hc.font=lf
            else:
                hc.font=df
            ws.cell(row=ri,column=5,value='').font=df  # Article/PR — analyst fills in
            ws.cell(row=ri,column=6,value='').font=df  # Sentiment — analyst fills in
            ws.cell(row=ri,column=7,value='').font=df  # Notes
            if art.get('is_paywall'):
                for ci in range(1,len(headers)+1):ws.cell(row=ri,column=ci).fill=pf
            for ci in range(1,len(headers)+1):ws.cell(row=ri,column=ci).border=bd
        ws.freeze_panes='A2'

    if is_trane_format:
        # TRANE FORMAT: Brand tabs + topic columns
        topics = ['Sustainability','Decarbonization','Innovation','Energy Efficiency','Digitization',
                  'Electrification','Workforce Development','Financial Performance','No Category Match']
        headers = ['Monitor Date','Date Published','Publication','Headline','Article OR Press Release',
                   'Sentiment'] + topics + ['CODED?','CEO/Executive Mention']
        ws_tot=wb.create_sheet("Totals")
        ws_tot['B2']='Honeywell';ws_tot['C2']='Carrier';ws_tot['D2']='JCI';ws_tot['E2']='Daikin';ws_tot['F2']='Lennox';ws_tot['G2']='Trane'
        for brand in brands:
            ws=wb.create_sheet(brand)
            for ci,h in enumerate(headers,1):c=ws.cell(row=1,column=ci,value=h);c.font=hf;c.fill=hfl;c.alignment=Alignment(horizontal='center',wrap_text=True);c.border=bd
            ws.column_dimensions['A'].width=12;ws.column_dimensions['B'].width=14;ws.column_dimensions['C'].width=22;ws.column_dimensions['D'].width=55;ws.column_dimensions['E'].width=18;ws.column_dimensions['F'].width=12
            for ci in range(7,16):ws.column_dimensions[chr(64+ci)].width=14
            ws.column_dimensions['P'].width=8;ws.column_dimensions['Q'].width=20
            ba = get_brand_articles_template(articles, brand)
            for ri,art in enumerate(ba,2):
                ws.cell(row=ri,column=1,value=art.get('monitor_date','')).font=df
                ws.cell(row=ri,column=2,value=art.get('date_published','')).font=df
                ws.cell(row=ri,column=3,value=art.get('publication','')).font=df
                hc=ws.cell(row=ri,column=4,value=art.get('headline','') or '(no headline)')
                url = art.get('url','')
                if url and art.get('headline','').strip():hc.hyperlink=url;hc.font=lf
                else:hc.font=df
                ws.cell(row=ri,column=5,value='').font=df
                if art.get('is_paywall'):
                    for ci in range(1,18):ws.cell(row=ri,column=ci).fill=pf
                for ci in range(1,18):ws.cell(row=ri,column=ci).border=bd
            ws.freeze_panes='A2'
        cm={'Honeywell':'B','Carrier':'C','JCI':'D','Daikin':'E','Lennox':'F','Trane Technologies':'G'}
        for b in brands:ws_tot[f'{cm.get(b,"B")}3']=len(get_brand_articles_template(articles,b))
        ws_def=wb.create_sheet("Key Topics & Definitions");ws_def['A1']='Key Topics';ws_def['B1']='Definitions'
        for di,(t,d) in enumerate([
            ('Sustainability','Climate change action'),('Decarbonization','Reducing carbon emissions'),
            ('Innovation','Advancing technologies'),('Energy Efficiency','Optimized energy use'),
            ('Digitization','Digital tech, AI, IoT'),('Electrification','Electrification of heat'),
            ('Workforce Development','Training, upskilling'),('Financial Performance','Revenue, earnings, stock'),
            ('No Category Match','Broader industry news')
        ],2):ws_def[f'A{di}']=t;ws_def[f'B{di}']=d
    else:
        # GENERIC FORMAT: One tab per document section, or single "All Articles" tab
        sections = {}
        for art in articles:
            sec = art.get('section', 'General')
            if sec not in sections: sections[sec] = []
            sections[sec].append(art)

        # Create a summary/totals sheet
        ws_tot = wb.create_sheet("Summary")
        ws_tot.cell(row=1, column=1, value="Section").font = hf
        ws_tot.cell(row=1, column=2, value="Articles").font = hf
        ws_tot.cell(row=1, column=1).fill = hfl
        ws_tot.cell(row=1, column=2).fill = hfl
        ws_tot.column_dimensions['A'].width = 35
        ws_tot.column_dimensions['B'].width = 12

        for si, (sec_name, sec_articles) in enumerate(sections.items()):
            # Summary row
            ws_tot.cell(row=si+2, column=1, value=sec_name).font = df
            ws_tot.cell(row=si+2, column=2, value=len(sec_articles)).font = df

            # Create tab for each section (truncate name to 31 chars for Excel limit)
            tab_name = sec_name[:31].replace('/', '-').replace('\\', '-').replace('*', '').replace('?', '').replace('[', '').replace(']', '')
            # Avoid duplicate tab names
            existing = [ws.title for ws in wb.worksheets]
            if tab_name in existing:
                tab_name = tab_name[:28] + f" ({si})"
            ws = wb.create_sheet(tab_name)
            _write_articles_to_sheet(ws, sec_articles)

        # Also create an "All Articles" tab with everything
        ws_all = wb.create_sheet("All Articles")
        _write_articles_to_sheet(ws_all, articles)

    out=BytesIO();wb.save(out);out.seek(0);return out


def get_brand_articles_template(articles, brand):
    """Distribute parsed articles to brand tabs based on section and keyword matching.
    Used for template mode (no AI tagging, so no brand.mentioned flags)."""
    result=[];bl=brand.lower()
    aliases={'trane technologies':['trane','metus','mitsubishi electric trane','thermo king'],
             'carrier':['carrier global','carrier'],'honeywell':['honeywell'],
             'jci':['johnson controls','jci'],'daikin':['daikin'],'lennox':['lennox']}
    ba=aliases.get(bl,[bl])
    for art in articles:
        sec=art.get('section','').lower()
        txt=(art.get('headline','')+' '+art.get('summary','')+' '+art.get('publication','')).lower()
        # Trane section articles go to Trane tab
        if bl=='trane technologies' and 'trane' in sec:
            result.append(art); continue
        # Benchmark and Trends articles: check if brand name or alias appears
        if 'benchmark' in sec or 'trends' in sec:
            for a in ba:
                if a in txt:result.append(art);break
    return result

def load_analyst_examples(ef):
    from openpyxl import load_workbook
    try:
        wb=load_workbook(BytesIO(ef.read()),read_only=True);exs=[]
        if 'Trane Technologies' in wb.sheetnames:
            ws=wb['Trane Technologies']
            tn=['Sustainability','Decarbonization','Innovation','Energy Efficiency','Digitization','Electrification','Workforce Development','Financial Performance','No Category Match']
            for ri,row in enumerate(ws.iter_rows(values_only=True)):
                if ri<2:continue
                if ri>12:break
                v=list(row)
                if len(v)>=16 and v[2]:
                    ts=[t for ti,t in enumerate(tn) if len(v)>6+ti and v[6+ti] and str(v[6+ti]).strip().lower()=='x']
                    exs.append(f"Pub:{v[2]}|HL:{str(v[3])[:60]}|Type:{v[4]}|Sent:{v[5]}|Topics:{','.join(ts) or 'NCM'}")
        wb.close();return '\n'.join(exs)
    except Exception:return ""


# ============================================================
# MAIN UI — SIDEBAR RESTORED + SPLIT SCREEN
# ============================================================
def _derive_project_name(filename):
    """Derive a short project name from the uploaded filename.
    'Daily Monitor Trane - 03.16.26.docx' → 'Trane'
    'Daily_News_Summary_Report_-_April_14__2026___.docx' → 'News_Summary_Report'
    """
    import os
    name = os.path.splitext(filename)[0]
    # Remove common prefixes
    for remove in ['Daily Monitor ', 'Daily_Monitor_', 'Daily_']:
        name = name.replace(remove, ' ')
    # Remove separators
    for remove in ['_-_', ' - ', '___', '__']:
        name = name.replace(remove, ' ')
    # Remove date patterns (numeric and word-based)
    name = re.sub(r'\d{2}[._]\d{2}[._]\d{2,4}', '', name)
    name = re.sub(r'(January|February|March|April|May|June|July|August|September|October|November|December)\s*\d{0,2}\s*,?\s*\d{0,4}', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\b\d{4}\b', '', name)
    # Clean up
    name = re.sub(r'[_\s]+', ' ', name).strip().strip('-').strip()
    if not name or len(name) < 2:
        name = os.path.splitext(filename)[0][:30]
    return name.replace(' ', '_')


def main():
    # Header
    st.markdown("""<div class="main-header">
        <div class="icon">🏷️</div>
        <div><h1>Media Intelligence Agent</h1>
        <p>Upload · Tag · Review with AI · Export</p></div>
    </div>""", unsafe_allow_html=True)

    # === SIDEBAR (always visible) ===
    with st.sidebar:
        st.markdown("### ⚙️ Configuration")

        # API Key
        saved_key = ""
        try: saved_key = st.secrets.get("ANTHROPIC_API_KEY", "")
        except Exception: pass
        if saved_key:
            api_key = saved_key
            st.success("✅ API Key loaded from secrets")
        else:
            api_key = st.text_input("Claude API Key", type="password", help="From console.anthropic.com")

        st.markdown("---")

        # How it works
        st.markdown("### 📋 How It Works")
        st.markdown("""
1. **Upload** Daily Monitor `.docx`
2. **Parser** extracts articles, URLs, summaries
3. **AI** reads articles & tags each one
4. **Review** in tile or table view
5. **Chat** with AI to correct tags
6. **Download** Excel with all brand tabs
        """)

        st.markdown("---")

        # Brands
        st.markdown("### 🔖 Brand Tabs")
        st.markdown("Trane Technologies (+ METUS, Thermo King) · Carrier · Honeywell · JCI · Daikin · Lennox")

        st.markdown("---")

        # Paywall blocklist
        st.markdown("### 🔒 Paywall Blocklist")
        st.caption(f"{len(PAYWALL_DOMAINS)} domains blocked")
        with st.expander("View / Edit blocklist"):
            st.markdown("Sites that will **never** be crawled — tagged from summary only.")
            bt = st.text_area("Domains (one per line)", value='\n'.join(sorted(PAYWALL_DOMAINS)), height=200)
            if st.button("Update blocklist"):
                PAYWALL_DOMAINS.clear()
                PAYWALL_DOMAINS.update({d.strip().lower() for d in bt.split('\n') if d.strip()})
                st.success(f"✅ {len(PAYWALL_DOMAINS)} domains")

        st.markdown("---")

        # Section filter (only when tagged)
        if st.session_state.get('tagged_articles'):
            st.markdown("### 🗂️ Filter by Section")
            sections = list(set(a.get('section','') for a in st.session_state.tagged_articles))
            selected_section = st.selectbox("Section", ["All Sections"] + sorted(sections))
            st.session_state['section_filter'] = selected_section

            st.markdown("---")
            st.markdown("### 📊 Quick Stats")
            tagged = st.session_state.tagged_articles
            st.markdown(f"**Total:** {len(tagged)} articles")
            st.markdown(f"**Paywalled:** {sum(1 for a in tagged if a.get('is_paywall'))}")
            st.markdown(f"**Low confidence:** {sum(1 for a in tagged if a.get('confidence',{}).get('overall')=='low')}")
            st.markdown(f"**CEO mentions:** {sum(1 for a in tagged if a.get('ceo_mention'))}")

    # === Initialize session state ===
    for key in ['chat_messages', 'tagged_articles', 'monitor_date', 'parsed_articles', 'section_filter', 'project_name']:
        if key not in st.session_state:
            st.session_state[key] = [] if key == 'chat_messages' else (None if key in ('tagged_articles','parsed_articles') else '')

    # ==========================================
    # PRE-TAGGING: Upload + Parse
    # ==========================================
    if st.session_state.tagged_articles is None:
        col1, col2 = st.columns([2, 1])
        with col1:
            uploaded_docx = st.file_uploader("📄 Upload Daily Monitor (.docx)", type=['docx'])
            if not uploaded_docx:
                st.markdown("""<div class="upload-panel">
                    <div class="icon">📄</div>
                    <div class="title">Drop your Daily Monitor here</div>
                    <div class="sub">.docx file — articles will be parsed automatically</div>
                </div>""", unsafe_allow_html=True)
        with col2:
            uploaded_excel = st.file_uploader("📊 Analyst Reference (optional)", type=['xlsx'])
            if not uploaded_excel:
                st.markdown("""<div class="upload-panel" style="padding:1.8rem;">
                    <div class="icon" style="font-size:2rem;">📊</div>
                    <div class="title" style="font-size:0.85rem;">Analyst reference</div>
                    <div class="sub">Helps AI learn your tagging patterns</div>
                </div>""", unsafe_allow_html=True)

        if uploaded_docx and st.session_state.parsed_articles is None:
            with st.spinner("🔍 Parsing Word document..."):
                fb = uploaded_docx.read()
                articles, md = parse_docx(fb, uploaded_docx.name)
            if not articles:
                st.error("No articles found."); return
            for art in articles:
                if art.get('url') and is_paywall_url(art['url']):
                    art['is_paywall'] = True; art['paywall_reason'] = 'blocklist'
                elif not art.get('url'):
                    art['is_paywall'] = True; art['paywall_reason'] = 'no_url'
            st.session_state.parsed_articles = articles
            st.session_state.monitor_date = md
            st.session_state.project_name = _derive_project_name(uploaded_docx.name)
            st.rerun()

        if st.session_state.parsed_articles is not None and st.session_state.tagged_articles is None:
            articles = st.session_state.parsed_articles
            md = st.session_state.monitor_date
            total = len(articles)
            exp = sum(1 for a in articles if a.get('is_expansion'))
            pay = sum(1 for a in articles if a.get('is_paywall'))
            crawl = total - pay

            st.markdown(f"""<div style="display:flex; flex-wrap:wrap; gap:0.4rem; margin:1rem 0;">
                <div class="stat-chip"><div class="stat-num">{total}</div><div class="stat-label">Articles</div></div>
                <div class="stat-chip"><div class="stat-num">{total-exp}</div><div class="stat-label">Main</div></div>
                <div class="stat-chip"><div class="stat-num">{exp}</div><div class="stat-label">Expansion</div></div>
                <div class="stat-chip"><div class="stat-num">{pay}</div><div class="stat-label">Paywalled</div></div>
                <div class="stat-chip"><div class="stat-num">{crawl}</div><div class="stat-label">Will Crawl</div></div>
                <div class="stat-chip"><div class="stat-num">{md or '—'}</div><div class="stat-label">Monitor Date</div></div>
            </div>""", unsafe_allow_html=True)

            if not api_key:
                st.warning("⚠️ Enter your Claude API Key in the sidebar to use AI Tagging.")

            st.markdown("---")
            st.markdown("**Choose your workflow:**")
            btn_col1, btn_col2 = st.columns(2)

            with btn_col1:
                st.markdown("""<div style="background:#f0fff4;border:1px solid #c6f6d5;border-radius:12px;padding:0.8rem;margin-bottom:0.5rem;">
                    <div style="font-weight:700;color:#276749;font-size:0.9rem;">📋 Template for Manual Tagging</div>
                    <div style="font-size:0.75rem;color:#4a5568;margin-top:0.3rem;">Parser fills columns A–D (Monitor Date, Date Published, Publication, Headline with hyperlinks). Columns E onwards are blank for analysts to tag manually.</div>
                </div>""", unsafe_allow_html=True)
                sd = md.replace(' ','_').replace('.','') if md else 'report'
                template_data = generate_template_excel(articles, md)
                st.download_button("📋 Download Template Excel", data=template_data,
                    file_name=f"{st.session_state.get('project_name','Report')}_Template_{sd}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)

            with btn_col2:
                st.markdown("""<div style="background:#ebf4ff;border:1px solid #bee3f8;border-radius:12px;padding:0.8rem;margin-bottom:0.5rem;">
                    <div style="font-weight:700;color:#3182ce;font-size:0.9rem;">🤖 AI Auto-Tagging</div>
                    <div style="font-size:0.75rem;color:#4a5568;margin-top:0.3rem;">AI reads each article, tags sentiment, topics, CEO mentions, article/PR type. Review and edit via chat agent. Requires API key.</div>
                </div>""", unsafe_allow_html=True)
                if api_key:
                    if st.button("🚀 Start AI Tagging", type="primary", use_container_width=True):
                        analyst_ex = load_analyst_examples(uploaded_excel) if uploaded_excel else ""
                        pbar = st.progress(0); stxt = st.empty(); larea = st.empty(); logs = []
                        def pcb(p, m):
                            pbar.progress(min(p,1.0))
                            stxt.markdown(f'<div style="font-size:0.85rem;font-weight:600;color:#667eea;padding:0.2rem 0;">{m}</div>', unsafe_allow_html=True)
                            logs.append(f"[{datetime.now().strftime('%H:%M:%S')}] {m}")
                            larea.markdown(f'<div style="background:#ffffff;border:1px solid #e8e9ee;border-radius:12px;padding:0.5rem 0.8rem;font-family:JetBrains Mono,monospace;font-size:0.7rem;color:#4a5568;white-space:pre-wrap;line-height:1.5;">{chr(10).join(logs[-8:])}</div>', unsafe_allow_html=True)
                        pcb(0.05, "Starting parallel processing...")
                        tagged = tag_articles_batch(articles, api_key, analyst_ex, pcb)
                        st.session_state.tagged_articles = tagged
                        low_conf = sum(1 for t in tagged if t.get('confidence',{}).get('overall') == 'low')
                        ok_count = sum(1 for t in tagged if not t.get('tags_failed'))
                        st.session_state.chat_messages = [{'role':'assistant','content':
                            f"✅ **{ok_count}/{len(tagged)}** articles tagged.\n\n"
                            + (f"⚠️ **{low_conf} articles** have low confidence. Ask: *'show low confidence articles'*\n\n" if low_conf > 0 else "")
                            + "**What you can do:**\n- Review articles (tiles/table) on the left\n- Chat here to correct: *'change #5 sentiment to Positive'*\n- Ask: *'why is article 8 tagged Decarbonization?'*\n- Export: *'generate Excel'*"}]
                        st.rerun()
                else:
                    st.info("Enter API key in sidebar to enable AI tagging.")
        return

    # ==========================================
    # POST-TAGGING: Status bar + Split Screen
    # ==========================================
    tagged = st.session_state.tagged_articles
    md = st.session_state.monitor_date
    brands = ['Trane Technologies','Carrier','Honeywell','JCI','Daikin','Lennox']
    topics = ['Sustainability','Decarbonization','Innovation','Energy Efficiency','Digitization',
              'Electrification','Workforce Development','Financial Performance','No Category Match']

    ok_count = sum(1 for t in tagged if not t.get('tags_failed'))
    low_c = sum(1 for t in tagged if t.get('confidence',{}).get('overall')=='low')
    med_c = sum(1 for t in tagged if t.get('confidence',{}).get('overall')=='medium')
    hi_c = sum(1 for t in tagged if t.get('confidence',{}).get('overall')=='high')
    pw_c = sum(1 for t in tagged if t.get('is_paywall'))
    sd = md.replace(' ','_').replace('.','') if md else 'report'

    dot = "status-dot-done" if ok_count == len(tagged) else "status-dot-error"
    st.markdown(f"""<div class="status-bar">
        <div class="status-dot {dot}"></div>
        <div class="status-text">Tagged {ok_count}/{len(tagged)} articles</div>
        <div class="status-detail">{hi_c} high · {med_c} med · {low_c} low · {pw_c} paywall · {md}</div>
    </div>""", unsafe_allow_html=True)

    # Download row
    dl1, dl2, dl3 = st.columns([2, 2, 6])
    with dl1:
        st.download_button("📥 Download Excel", data=generate_excel(tagged,md),
            file_name=f"{st.session_state.get('project_name','Report')}_Tagged_{sd}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary", use_container_width=True)
    with dl2:
        st.download_button("📥 Download CSV", data=generate_csv(tagged,topics),
            file_name=f"{st.session_state.get('project_name','Report')}_Tagged_{sd}.csv", mime="text/csv", use_container_width=True)

    # Split: 60% results, 40% chat
    left_col, right_col = st.columns([60, 40])

    with left_col:
        vc1, vc2 = st.columns([3,1])
        with vc2:
            view = st.radio("View", ["Tiles","Table"], horizontal=True, label_visibility="collapsed")

        # Apply section filter from sidebar
        section_filter = st.session_state.get('section_filter', 'All Sections')
        display_tagged = tagged if section_filter == 'All Sections' else [a for a in tagged if a.get('section') == section_filter]

        tabs = st.tabs(["All"] + brands)
        with tabs[0]:
            if view == "Tiles": tile_view(display_tagged, topics, "all")
            else: table_view(display_tagged, topics, "all")
        for bi, b in enumerate(brands):
            with tabs[bi+1]:
                ba = get_brand_articles(display_tagged, b)
                if ba:
                    if view == "Tiles": tile_view(ba, topics, b)
                    else: table_view(ba, topics, b)
                else: st.caption(f"No articles for {b}")

    with right_col:
        st.markdown("""<div class="chat-header">
            <div class="dot"></div>
            <div><h3>Tagging Agent</h3><span>Online — ask anything about the articles</span></div>
        </div>""", unsafe_allow_html=True)

        chat_area = st.container(height=500)
        with chat_area:
            for msg in st.session_state.chat_messages:
                with st.chat_message(msg['role'], avatar="🏷️" if msg['role']=='assistant' else "👤"):
                    st.markdown(msg['content'])

        if prompt := st.chat_input("e.g. 'show low confidence' or 'change #5 to Positive'"):
            st.session_state.chat_messages.append({'role':'user','content':prompt})
            with chat_area:
                with st.chat_message("user", avatar="👤"): st.markdown(prompt)
                with st.chat_message("assistant", avatar="🏷️"):
                    with st.spinner("Thinking..."):
                        response = process_chat_command(prompt, tagged, api_key)
                        changes = apply_chat_changes(tagged, response)
                        if changes == -1:
                            st.markdown("Generating report...")
                            st.download_button("📥 Download Excel", data=generate_excel(tagged,md), file_name=f"{st.session_state.get('project_name','Report')}_Tagged_{sd}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                            rc = "📊 Report ready — click download above."
                        elif changes > 0:
                            rc = re.sub(r'\{[^}]*"action"[^}]*\}','',response,flags=re.DOTALL).strip()
                            rc += f"\n\n✅ **{changes} change(s) applied.**"
                            st.session_state.tagged_articles = tagged
                        else: rc = response
                        st.markdown(rc)
                        st.session_state.chat_messages.append({'role':'assistant','content':rc})
            st.rerun()

        st.caption("Try: *'show low confidence'* · *'change #5 to Positive'* · *'generate Excel'*")


# ============================================================
# TILE VIEW WITH PAYWALL HIGHLIGHTING
# ============================================================
def tile_view(articles, topics, brand_filter):
    if not articles: st.caption("No articles."); return
    for i, art in enumerate(articles):
        bd=art.get('brands',{});td=art.get('topics',{});conf=art.get('confidence',{})
        if brand_filter=="all":
            sent="Neutral"
            for b in ['Trane Technologies','Carrier','Honeywell','JCI','Daikin','Lennox']:
                if bd.get(b,{}).get('mentioned'):sent=bd[b].get('sentiment','Neutral');break
        else: sent=bd.get(brand_filter,{}).get('sentiment','Neutral')
        active=[t for t in topics if td.get(t,False)]
        sc='tag-positive' if sent=='Positive' else ('tag-negative' if sent=='Negative' else 'tag-neutral')
        hl=art.get('headline','No headline');url=art.get('url','')
        hl_html=f'<a href="{url}" target="_blank">{hl}</a>' if url else hl

        is_pw = art.get('is_paywall', False)
        card_class = "article-card article-card-paywall" if is_pw else "article-card"

        badges=""
        if is_pw: badges+='<span class="paywall-banner">🔒 Tagged from summary only — needs paywall access for full analysis</span><br>'
        if art.get('is_translated'):badges+='<span class="tag-chip tag-translated">🌐 Translated</span>'
        if art.get('is_expansion'):badges+='<span class="tag-chip tag-expansion">🔗 Expansion</span>'
        if art.get('ceo_mention'):badges+=f'<span class="tag-chip tag-ceo">{art["ceo_mention"]}</span>'
        if art.get('tags_failed'):badges+='<span class="tag-chip tag-low">⚠️ Fallback</span>'
        if art.get('publication_flagged'):badges+='<span class="tag-chip tag-paywall">⚠️ Check Pub Name</span>'

        overall_conf = conf.get('overall','')
        if overall_conf == 'high': badges += '<span class="tag-chip tag-high">✓ High</span>'
        elif overall_conf == 'medium': badges += '<span class="tag-chip tag-medium">~ Medium</span>'
        elif overall_conf == 'low': badges += '<span class="tag-chip tag-low">⚠ Low</span>'

        tchips=''.join(f'<span class="tag-chip tag-topic">{t}</span>' for t in active) or '<span class="tag-chip tag-neutral">No Category Match</span>'
        try: idx = st.session_state.tagged_articles.index(art) + 1
        except: idx = i + 1

        st.markdown(f"""<div class="{card_class}">
            <div class="card-num">#{idx}</div>
            <div class="pub-name">{art.get('publication','')} · {art.get('date_published','')} · <span class="{sc}" style="padding:0.1rem 0.4rem;border-radius:5px;font-size:0.62rem;">{sent}</span> · {art.get('article_or_pr','Article')}</div>
            {f'<div>{badges}</div>' if badges else ''}
            <div class="headline">{hl_html}</div>
            <div class="meta">{tchips}</div>
        </div>""", unsafe_allow_html=True)
    st.caption(f"{len(articles)} articles — use # in chat to edit")


def table_view(articles, topics, brand_filter):
    import pandas as pd
    if not articles: st.caption("No articles."); return
    rows=[]
    for i,art in enumerate(articles):
        bd=art.get('brands',{});td=art.get('topics',{});conf=art.get('confidence',{})
        if brand_filter=="all":
            sent="—"
            for b in ['Trane Technologies','Carrier','Honeywell','JCI','Daikin','Lennox']:
                if bd.get(b,{}).get('mentioned'):sent=bd[b].get('sentiment','Neutral');break
        else: sent=bd.get(brand_filter,{}).get('sentiment','Neutral')
        try: idx = st.session_state.tagged_articles.index(art)+1
        except: idx = i+1
        row={'#':idx,'Publication':art.get('publication',''),'Headline':art.get('headline','')[:60],
             'Type':art.get('article_or_pr','Article'),'Sentiment':sent,'Conf':conf.get('overall','?')}
        for t in topics:row[t]='x' if td.get(t,False) else ''
        row['CEO']=art.get('ceo_mention','')
        flags=[]
        if art.get('is_paywall'):flags.append('🔒 PW')
        if art.get('is_translated'):flags.append('🌐')
        if art.get('tags_failed'):flags.append('⚠️')
        row['Flags']=' '.join(flags)
        rows.append(row)
    st.dataframe(pd.DataFrame(rows),use_container_width=True,height=min(len(rows)*38+40,600))
    st.caption(f"{len(rows)} articles | 🔒 = Paywall (summary only)")


if __name__ == "__main__":
    main()
